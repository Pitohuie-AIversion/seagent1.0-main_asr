import openpyxl
import yaml
from pathlib import Path

excel_path = Path("/root/mzy/seagent1.0-main_asr/config/SEAgent_ROV_AUV_Payload能力矩阵_括号清理版.xlsx")
fleet_path = Path("/root/mzy/seagent1.0-main_asr/config/robot_fleet.yaml")

# 1. Parse Sheet 01 from Excel
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['01_机器人展示']

robots_excel = []
curr = None
for r in range(1, sheet.max_row + 1):
    v1 = sheet.cell(r, 1).value
    v5 = sheet.cell(r, 5).value
    if v1 and (' Onboard ' in str(v1) or '｜' in str(v1)):
        if str(v1).startswith('ROV / AUV'):
            continue
        if curr:
            robots_excel.append(curr)
        curr = {'title': str(v1).strip(), 'onboard': [], 'supported': []}
        continue
    if curr:
        if v1 and v1 not in ['Onboard', '已清理：独立高清摄像机 / LED照明及历史错误项', '黄色：新增 / 统一名称'] and not str(v1).startswith('ROV / AUV'):
            curr['onboard'].append(str(v1).strip())
        if v5 and v5 not in ['Supported', '粉色：删除 / 清理说明']:
            curr['supported'].append(str(v5).strip())
if curr:
    robots_excel.append(curr)

print(f"Extracted {len(robots_excel)} robots from Sheet 01:")
for i, bot in enumerate(robots_excel):
    print(f"Robot {i+1}: {bot['title']} (Onboard={len(bot['onboard'])}, Supported={len(bot['supported'])})")

# 2. Map robots in Sheet 01 to variant keys in robot_fleet.yaml
variant_keys = [
    "crawler_heavy_seabed_robot_1600hp",
    "special_work_class_robot_600hp",
    "general_work_class_rov_250hp",
    "light_work_class_rov_150hp",
    "observation_rov_75hp",
    "autonomous_underwater_vehicle_324cc"
]

variant_payload_map = {}
for i, vk in enumerate(variant_keys):
    variant_payload_map[vk] = {
        "onboard_payloads": robots_excel[i]["onboard"],
        "supported_payloads": robots_excel[i]["supported"]
    }

# 3. Read existing robot_fleet.yaml lines and replace payload blocks for each variant
with open(fleet_path, "r", encoding="utf-8") as f:
    lines = f.readlines()

new_lines = []
current_variant = None
in_onboard = False
in_supported = False
skip_mode = False

i = 0
while i < len(lines):
    line = lines[i]
    stripped = line.strip()
    
    # Track current model_variant key
    if line.startswith("  ") and not line.startswith("    ") and stripped.endswith(":") and not stripped.startswith("#"):
        key = stripped[:-1].strip()
        if key in variant_payload_map:
            current_variant = key
        elif current_variant and not key.startswith("hard_params"):
            # Exited variant block if top level indentation matches
            pass

    if current_variant and "onboard_payloads:" in line and not stripped.startswith("#"):
        # Replace onboard_payloads block
        new_lines.append("      onboard_payloads:\n")
        for item in variant_payload_map[current_variant]["onboard_payloads"]:
            new_lines.append(f'        - "{item}"\n')
        # Skip original onboard_payloads items
        i += 1
        while i < len(lines) and lines[i].startswith("        - ") and not lines[i].strip().startswith("#"):
            i += 1
        continue

    if current_variant and "supported_payloads:" in line and not stripped.startswith("#"):
        # Replace supported_payloads block
        new_lines.append("      supported_payloads:\n")
        for item in variant_payload_map[current_variant]["supported_payloads"]:
            new_lines.append(f'        - "{item}"\n')
        # Skip original supported_payloads items
        i += 1
        while i < len(lines) and lines[i].startswith("        - ") and not lines[i].strip().startswith("#"):
            i += 1
        continue

    new_lines.append(line)
    i += 1

with open(fleet_path, "w", encoding="utf-8") as f:
    f.writelines(new_lines)

print(f"\nSuccessfully replaced onboard and supported payloads in {fleet_path}!")
