import openpyxl
import yaml

excel_path = '/root/mzy/seagent1.0-main_asr/config/SEAgent_ROV_AUV_Payload能力矩阵_括号清理版.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Parse Sheet 02
s2 = wb['02_工作类型展示']
worktypes = {}
curr = None
for r in range(1, s2.max_row + 1):
    c1 = s2.cell(row=r, column=1).value
    if c1 and '｜' in str(c1) and '项' in str(c1):
        curr = str(c1).split('｜')[0].strip()
        worktypes[curr] = []
    elif curr and c1:
        t = str(c1).strip()
        if t not in ['支持工具', '黄色：新增 / 统一名称'] and not t.startswith('口径：') and not t.startswith('已统一'):
            worktypes[curr].append(t)

print("=== SHEET 02 WORKTYPES ===")
all_tools = set()
for wt, tlist in worktypes.items():
    print(f"{wt} ({len(tlist)} items):")
    for t in tlist:
        print(f"  - {t}")
    all_tools.update(tlist)

print(f"\nTotal unique tools across Sheet 02: {len(all_tools)}")

# Read assets.yaml
assets_path = '/root/mzy/seagent1.0-main_asr/config/assets.yaml'
with open(assets_path, 'r', encoding='utf-8') as f:
    assets = yaml.safe_load(f)

catalog = assets.get('payload_catalog', {})
catalog_names = set()
catalog_map = {}
for k, v in catalog.items():
    if isinstance(v, dict):
        if 'name' in v:
            catalog_names.add(v['name'])
            catalog_map[v['name']] = k
        if 'aliases' in v and isinstance(v['aliases'], list):
            for a in v['aliases']:
                catalog_names.add(a)
                catalog_map[a] = k

missing_tools = sorted([t for t in all_tools if t not in catalog_names])
print(f"\nTools in Sheet 02 missing from payload_catalog ({len(missing_tools)}):")
for t in missing_tools:
    print(" -", t)

# Also check Sheet 05 & Sheet 06 info
print("\n=== SHEET 05 ITEMS ===")
s5 = wb['05_机械臂工具与水质传感器']
for r in range(1, s5.max_row + 1):
    vals = [s5.cell(row=r, column=c).value for c in range(1, s5.max_column + 1)]
    if any(vals):
        print(f"Row {r:2d}: {vals}")

print("\n=== SHEET 06 ITEMS ===")
s6 = wb['06_重量依据']
for r in range(1, s6.max_row + 1):
    vals = [s6.cell(row=r, column=c).value for c in range(1, s6.max_column + 1)]
    if any(vals):
        print(f"Row {r:2d}: {vals}")
