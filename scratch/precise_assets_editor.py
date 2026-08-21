import openpyxl

excel_path = '/root/mzy/seagent1.0-main_asr/config/SEAgent_ROV_AUV_Payload能力矩阵_括号清理版.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

s2 = wb['02_工作类型展示']
worktypes = {}
curr = None
for r in range(1, s2.max_row + 1):
    c1 = s2.cell(row=r, column=1).value
    if c1 and '｜' in str(c1) and '项' in str(c1):
        curr = str(c1).split('｜')[0].strip()
        worktypes[curr] = []
    elif curr and c1:
        t = str(c1).strip()
        if t not in ['支持工具', '黄色：新增 / 统一名称'] and not t.startswith('口径：') and not t.startswith('已统一'):
            worktypes[curr].append(t)

assets_path = '/root/mzy/seagent1.0-main_asr/config/assets.yaml'

with open(assets_path, 'r', encoding='utf-8') as f:
    content = f.read()

# Generate new common block formatted strings
def format_common(tools):
    lines = ["    common:"]
    for t in tools:
        lines.append(f'      - "{t}"')
    return "\n".join(lines)

insp_common = format_common(worktypes['管缆巡检'])
burial_common = format_common(worktypes['管缆埋设'])
tree_common = format_common(worktypes['采油树控制面板插拔操作'])

import re

# 1. Replace pipeline_inspection common block
content = re.sub(
    r'(pipeline_inspection:\n\s*)common:.*?(?=\n\s*description:)',
    r'\g<1>' + insp_common.strip(),
    content,
    flags=re.DOTALL
)

# 2. Replace pipeline_burial common block
content = re.sub(
    r'(pipeline_burial:\n\s*)common:.*?(?=\n\s*description:)',
    r'\g<1>' + burial_common.strip(),
    content,
    flags=re.DOTALL
)

# 3. Replace tree_valve_operation common block
content = re.sub(
    r'(tree_valve_operation:\n\s*)common:.*?(?=\n\s*description:)',
    r'\g<1>' + tree_common.strip(),
    content,
    flags=re.DOTALL
)

# 4. Add aliases to existing catalog items in text
alias_additions = {
    'forward_looking_sonar': ['FLS声呐系统', 'FLS声呐', 'FLS'],
    'dvl': ['DVL测速系统'],
    'usbl': ['USBL定位系统'],
    'chassis_skid_module': ['履带模块'],
    'cleaning_brush': ['旋转清洗刷'],
    'gripper': ['机械臂末端夹爪 / 夹具', '机械臂末端夹爪', '末端夹爪'],
    'hydraulic_cutter': ['液压剪切 / 水下锯切工具', '液压电缆剪切器'],
    'tss_cable_tracker': ['TSS管缆跟踪系统']
}

for item_key, new_aliases in alias_additions.items():
    pattern = rf'({item_key}:.*?\naliases:\s*\[)(.*?)(\])'
    def add_alias(m):
        prefix, existing, suffix = m.group(1), m.group(2), m.group(3)
        existing_items = [x.strip().strip('"\'') for x in existing.split(',')]
        for a in new_aliases:
            if a not in existing_items:
                existing_items.append(a)
        formatted = ", ".join([f'"{x}"' for x in existing_items])
        return f'{prefix}{formatted}{suffix}'
    content = re.sub(pattern, add_alias, content, flags=re.DOTALL)

# 5. Append new catalog items at the end of assets.yaml
new_catalog_yaml = """
  turbid_water_camera:
    name: "浑水摄像机"
    aliases: ["浑水摄像机", "浑水相机", "OrphieCam", "浊度透视摄像机"]
    description: "专用于低能见度和高浑浊度水下环境的强穿透高清晰度光学观察与巡检摄像机。"
    capabilities: ["低能见度巡检", "高浑浊度光学成像", "水下观测"]
    weight_in_air_kg: 4.5 # 空气中质量(kg)
    weight_in_water_kg: null

  independent_sampler:
    name: "独立采样器"
    aliases: ["独立采样器", "车载独立采样器", "水下独立采样装置"]
    description: "非机械臂依赖的独立式深海水体与沉积物/环境DNA采样装置。"
    capabilities: ["独立水体采样", "环境DNA采样", "物理化学采样"]
    weight_in_air_kg: 15.0 # 空气中质量(kg)
    weight_in_water_kg: null

  valve_torque_tool:
    name: "阀门扭矩工具"
    aliases: ["阀门扭矩工具", "Torque Tool", "阀门操作工具", "Class 1-4 Torque Tool"]
    description: "符合 ISO 13628-8 / API 17H 标准的工作级 ROV 阀门扭矩操作工具，用于水下采油树控制面板及管道阀门的高扭矩旋转开关操作。"
    capabilities: ["阀门扭矩旋转", "ISO接口操作", "高扭矩开关"]
    weight_in_air_kg: 45.0 # 空气中质量(kg)
    weight_in_water_kg: 35.0 # 水下质量(kg)

  rov_hydraulic_hot_stab:
    name: "ROV液压热插拔接头"
    aliases: ["ROV液压热插拔接头", "Hot Stab", "液压热插拔", "水下热插拔接头"]
    description: "ROV 水下液压热插拔接头，用于在水下压力环境下进行流体注液、液压回路高压连接及临时控制通道对接。"
    capabilities: ["水下液压热插拔", "高压流体连接", "临时液压控制"]
    weight_in_air_kg: 4.6 # 空气中质量(kg)
    weight_in_water_kg: null

  hydraulic_flying_lead_tool:
    name: "液压飞线插拔工具"
    aliases: ["液压飞线插拔工具", "FLOT", "飞线工具", "Flying Lead Orientation Tool"]
    description: "液压飞线定向与对准插拔工具（FLOT），用于辅助 ROV 进行水下液压飞线的精准对准、旋转锁紧与快捷插拔操作。"
    capabilities: ["液压飞线对准", "飞线旋转插拔", "水下连接干预"]
    weight_in_air_kg: 45.0 # 空气中质量(kg)
    weight_in_water_kg: 25.0 # 水下质量(kg)

  electric_flying_lead_tool:
    name: "电气飞线插拔工具"
    aliases: ["电气飞线插拔工具", "EFL插拔工具", "电气飞线工具", "EFL Connector"]
    description: "用于水下控制系统电气飞线（EFL）的专用或机械臂直接插拔操作工具，实现水下信号与电力电缆的湿插拔连接。"
    capabilities: ["电气飞线湿插拔", "信号与电力连接", "水下接头安装"]
    weight_in_air_kg: 5.0 # 空气中质量(kg)
    weight_in_water_kg: null

  six_axis_force_torque_sensor:
    name: "六维力 / 力矩传感器"
    aliases: ["六维力 / 力矩传感器", "六维力传感器", "力矩传感器", "6轴力传感器"]
    description: "安装于机械臂腕部或末端工具接口，实时感知三维空间内的力与力矩大小，为机械臂柔顺控制与精准触觉反馈提供保障。"
    capabilities: ["六维力控反馈", "触觉感知", "柔顺干预控制"]
    weight_in_air_kg: 2.0 # 空气中质量(kg)
    weight_in_water_kg: null

  manipulator_wrist_camera:
    name: "机械臂腕部摄像机"
    aliases: ["机械臂腕部摄像机", "腕部摄像机", "腕部相机", "Wrist Camera"]
    description: "紧凑型深水高清摄像机，直接安装于机械臂腕部，为末端夹爪工具对准、热插拔及精细微操作提供近距离视角。"
    capabilities: ["末端对准视角", "微操作监控", "近距离光学观察"]
    weight_in_air_kg: 0.7 # 空气中质量(kg)
    weight_in_water_kg: null

  underwater_imaging_system:
    name: "水下成像系统"
    aliases: ["水下成像系统", "成像系统", "光学成像系统"]
    description: "综合水下成像系统，集成高分辨率摄像机、补光灯与画面实时处理模块，用于全过程视觉观察与作业记录。"
    capabilities: ["全过程视觉观察", "作业画面记录", "水下图像感知"]
    weight_in_air_kg: null
    weight_in_water_kg: null

  five_dof_manipulator:
    name: "五自由度机械臂"
    aliases: ["五自由度机械臂", "5自由度机械臂", "5自由度抓持臂"]
    description: "5自由度重型抓持或辅助机械臂，主要用于稳定贴靠构件、辅助持物及配合主臂进行水下作业。"
    capabilities: ["辅助抓持", "构件固定", "辅助干预"]
    weight_in_air_kg: null
    weight_in_water_kg: null

  manipulator_tool_quick_change:
    name: "机械臂工具快换装置"
    aliases: ["机械臂工具快换装置", "工具快换装置", "快换接头"]
    description: "用于机械臂末端快速切换不同作业工具的快捷机械/液压锁紧接口。"
    capabilities: ["末端工具快换", "水下工装切换", "多功能扩展"]
    weight_in_air_kg: null
    weight_in_water_kg: null

  manipulator_sampling_tool:
    name: "机械臂相关采样工具"
    aliases: ["机械臂相关采样工具", "机械臂采样工具", "机械臂采样工装"]
    description: "配合机械臂使用的推管采样器、抓斗或防污染密封箱，用于采集深海沉积物、岩石或生物样品。"
    capabilities: ["机械臂辅助采样", "沉积物/样品采集", "防污染封装"]
    weight_in_air_kg: null
    weight_in_water_kg: null

  cable_depressor:
    name: "压缆器"
    aliases: ["压缆器", "缆压器", "海缆压块"]
    description: "配合管缆埋设机使用的压缆引导装置，用于在喷冲开沟后将海底管缆顺畅压入泥面以下并保持埋设姿态。"
    capabilities: ["管缆压入泥面", "埋设姿态控制", "海缆引导"]
    weight_in_air_kg: 1000.0 # 空气中质量(kg)
    weight_in_water_kg: null

  burial_depth_sensor:
    name: "埋深感应装置"
    aliases: ["埋深感应装置", "埋深传感器", "埋深检测器"]
    description: "用于实时感应管缆埋入泥面以下深度的电磁或机械触感测头，配合埋深控制系统进行闭环调节。"
    capabilities: ["埋深实时感应", "闭环反馈", "路由埋深测量"]
    weight_in_air_kg: 150.0 # 空气中质量(kg)
    weight_in_water_kg: null

  usbl_auxiliary_positioning_module:
    name: "USBL辅助定位模块"
    aliases: ["USBL辅助定位模块", "USBL辅助定位", "AUV定位辅助模块"]
    description: "专用于 AUV 或小型 ROV 的超短基线辅助声学定位与应答模块，提升定位精度与数据传输效率。"
    capabilities: ["AUV辅助定位", "声学坐标修正", "水下定位导航"]
    weight_in_air_kg: null
    weight_in_water_kg: null

  dissolved_oxygen_sensor:
    name: "溶解氧传感器"
    aliases: ["溶解氧传感器", "DO传感器", "光学溶解氧传感器"]
    description: "用于精确测量水体中溶解氧浓度与饱和度的光学/电化学传感器，广泛应用于海洋生态与环境监测。"
    capabilities: ["溶解氧测量", "海洋水质监测", "生态环境评估"]
    weight_in_air_kg: null
    weight_in_water_kg: null

  turbidity_sensor:
    name: "浊度传感器"
    aliases: ["浊度传感器", "光学浊度传感器", "浊度计"]
    description: "利用光学散射原理测量水体悬浮物浓度与浊度的传感器，用于监测海底施工羽流与泥沙扰动。"
    capabilities: ["浊度监测", "施工羽流跟踪", "悬浮物浓度测量"]
    weight_in_air_kg: null
    weight_in_water_kg: null
"""

# Append new_catalog_yaml to content if not already present
if 'turbid_water_camera:' not in content:
    content = content.rstrip() + "\n" + new_catalog_yaml

with open(assets_path, 'w', encoding='utf-8') as f:
    f.write(content)

print("Targeted precise assets update complete!")
