import yaml
import openpyxl

excel_path = '/root/mzy/seagent1.0-main_asr/config/SEAgent_ROV_AUV_Payload能力矩阵_括号清理版.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

s2 = wb['02_工作类型展示']
worktypes = {}
curr = None
for r in range(1, s2.max_row + 1):
    c1 = s2.cell(row=r, column=1).value
    if c1 and '｜' in str(c1) and '项' in str(c1):
        curr = str(c1).split('｜')[0].strip()
        worktypes[curr] = []
    elif curr and c1:
        t = str(c1).strip()
        if t not in ['支持工具', '黄色：新增 / 统一名称'] and not t.startswith('口径：') and not t.startswith('已统一'):
            worktypes[curr].append(t)

assets_path = '/root/mzy/seagent1.0-main_asr/config/assets.yaml'

with open(assets_path, 'r', encoding='utf-8') as f:
    assets = yaml.safe_load(f)

# Update payload_options
assets['payload_options']['pipeline_inspection']['common'] = worktypes['管缆巡检']
assets['payload_options']['pipeline_burial']['common'] = worktypes['管缆埋设']
assets['payload_options']['tree_valve_operation']['common'] = worktypes['采油树控制面板插拔操作']

catalog = assets['payload_catalog']

# Remove '浊度传感器' and '溶解氧传感器' from water_quality_sensor_suite aliases to avoid alias collision
if 'water_quality_sensor_suite' in catalog:
    w_aliases = catalog['water_quality_sensor_suite'].get('aliases', [])
    w_aliases = [a for a in w_aliases if a not in ['浊度传感器', '溶解氧传感器']]
    catalog['water_quality_sensor_suite']['aliases'] = w_aliases

# Ensure every tool name in Sheet 02 is either the primary `name` of an existing catalog entry OR a new catalog entry
# Map of sheet tool name -> catalog key (updating existing key name or creating new key)
sheet_tool_to_key = {
    'FLS声呐系统': 'forward_looking_sonar',
    'DVL测速系统': 'dvl',
    'USBL定位系统': 'usbl',
    '履带模块': 'chassis_skid_module',
    '旋转清洗刷': 'cleaning_brush',
    '机械臂末端夹爪 / 夹具': 'gripper',
    '液压剪切 / 水下锯切工具': 'hydraulic_cutter',
    'TSS管缆跟踪系统': 'tss_cable_tracker',
    '多波束测深声呐': 'multibeam_sonar',
    '机械扫描声呐': 'mechanical_scanning_sonar',
    '高清水下摄像机': 'hd_camera',
    '云台摄像机': 'pan_tilt_camera',
    'LED水下照明灯': 'led_lighting',
    '激光标尺': 'laser_scaling',
    '双目视觉模块': 'stereo_vision_module',
    '前视声呐': 'forward_looking_sonar',
    '成像声呐': 'imaging_sonar',
    '侧扫声呐': 'side_scan_sonar',
    '浅地层剖面仪': 'sub_bottom_profiler',
    '腐蚀检测探头': 'corrosion_probe',
    '厚度检测传感器': 'thickness_sensor',
    '泄漏检测传感器': 'leak_detector',
    'TSS管缆跟踪传感器': 'tss_cable_tracker',
    '管缆外观检测设备': 'visual_inspection_skid',
    '三维视觉系统': 'three_dimensional_vision',
    '水质传感器': 'water_quality_sensor_suite',
    '合成孔径声呐': 'synthetic_aperture_sonar',
    'CTD': 'ctd_sensor',
    '高压水射流喷冲埋设模块': 'high_pressure_water_jet',
    '水下定位信标': 'acoustic_beacon',
    '海床地质探测设备': 'seabed_geology_detector',
    '机械切割开沟模块': 'mechanical_trenching_module',
    '海缆压埋装置': 'cable_burial_device',
    '埋深控制装置': 'burial_depth_controller',
    '声学应答器': 'acoustic_beacon',
    '海缆保护施工工具': 'cable_protection_tool',
    '多功能液压机械臂': 'multi_function_hydraulic_arm',
    '电液机械臂': 'electro_hydraulic_arm',
    '轻型多功能液压机械臂': 'multi_function_hydraulic_arm',
    '轻型电液机械臂': 'electro_hydraulic_arm'
}

# For entries where sheet_tool_name should become primary `name` or added as separate entry if name differs
# Let's check catalog_names
catalog_names = {item.get('name') for item in catalog.values()}

all_sheet_tools = set(worktypes['管缆巡检'] + worktypes['管缆埋设'] + worktypes['采油树控制面板插拔操作'])

# Check which tools in all_sheet_tools are not primary `name` of any catalog item
for tool_name in all_sheet_tools:
    if tool_name not in catalog_names:
        # Check if we can rename or create a key
        if tool_name in sheet_tool_to_key:
            target_key = sheet_tool_to_key[tool_name]
            if target_key in catalog:
                old_name = catalog[target_key]['name']
                # If old_name is different, make old_name an alias and set new name
                if old_name != tool_name:
                    aliases = catalog[target_key].get('aliases', [])
                    if old_name not in aliases:
                        aliases.append(old_name)
                    catalog[target_key]['name'] = tool_name
                    catalog[target_key]['aliases'] = aliases
                    catalog_names.add(tool_name)

print("Check missing tools after renaming/mapping existing keys:")
catalog_names = {item.get('name') for item in catalog.values()}
missing = [t for t in sorted(list(all_sheet_tools)) if t not in catalog_names]
print("Missing:", missing)
