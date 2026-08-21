import yaml
import openpyxl

excel_path = '/root/mzy/seagent1.0-main_asr/config/SEAgent_ROV_AUV_Payload能力矩阵_括号清理版.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

s2 = wb['02_工作类型展示']
worktypes = {}
curr = None
for r in range(1, s2.max_row + 1):
    c1 = s2.cell(row=r, column=1).value
    if c1 and '｜' in str(c1) and '项' in str(c1):
        curr = str(c1).split('｜')[0].strip()
        worktypes[curr] = []
    elif curr and c1:
        t = str(c1).strip()
        if t not in ['支持工具', '黄色：新增 / 统一名称'] and not t.startswith('口径：') and not t.startswith('已统一'):
            worktypes[curr].append(t)

assets_path = '/root/mzy/seagent1.0-main_asr/config/assets.yaml'

with open(assets_path, 'r', encoding='utf-8') as f:
    assets = yaml.safe_load(f)

# Update payload_options
assets['payload_options']['pipeline_inspection']['common'] = worktypes['管缆巡检']
assets['payload_options']['pipeline_burial']['common'] = worktypes['管缆埋设']
assets['payload_options']['tree_valve_operation']['common'] = worktypes['采油树控制面板插拔操作']

catalog = assets['payload_catalog']

# Restore original primary names for items expected by tests
catalog['multibeam_sonar']['name'] = "多波束声呐"
catalog['mechanical_scanning_sonar']['name'] = "机械式声呐"
catalog['tss_cable_tracker']['name'] = "TSS管缆跟踪传感器"
catalog['forward_looking_sonar']['name'] = "前视声呐"
catalog['dvl']['name'] = "DVL多普勒测速仪"
catalog['usbl']['name'] = "USBL定位设备"
catalog['chassis_skid_module']['name'] = "履带与滑橇组件"
catalog['cleaning_brush']['name'] = "清洗刷"
catalog['gripper']['name'] = "夹爪"
catalog['hydraulic_cutter']['name'] = "液压剪切器"

# Define dedicated catalog entries for Sheet 02 tool names that differ slightly from legacy catalog primary names
distinct_sheet_entries = {
    'multibeam_depth_sonar': {
        'name': '多波束测深声呐',
        'aliases': ['多波束测深声纳'],
        'description': '通过多个声学波束对海床和水下目标进行高精度条带式测深与三维地形测绘的多波束声呐系统。',
        'capabilities': ['海床测绘', '地形测量', '多波束测深']
    },
    'mechanical_scanning_sonar_sys': {
        'name': '机械扫描声呐',
        'aliases': ['机械扫描声纳'],
        'description': '利用机械换能器旋转扫描水下环境的声学成像设备，用于周侧 360 度远距离目标识别与声学成像。',
        'capabilities': ['环视声学成像', '周侧障碍物扫描', '水下目标搜寻']
    },
    'tss_cable_tracking_system': {
        'name': 'TSS管缆跟踪系统',
        'aliases': ['TSS管缆跟踪仪', 'TSS系统'],
        'description': '利用电磁感应原理探测埋入海床底质下方的金属管缆或通电电缆，实时测量管缆偏航与埋深的感知系统。',
        'capabilities': ['埋深测量', '电磁管缆跟踪', '路由定位']
    },
    'fls_sonar_system': {
        'name': 'FLS声呐系统',
        'aliases': ['FLS声呐', 'FLS'],
        'description': '用于机器人前方水域远距离障碍物探测、地形识别与水下碰撞预警的前视声学感知系统。',
        'capabilities': ['水下避障', '前方地形探测', '导航辅助']
    },
    'dvl_speed_system': {
        'name': 'DVL测速系统',
        'aliases': ['DVL测速'],
        'description': '通过声学多普勒效应测量水下机器人相对海床或水体三维运动速度的多普勒测速系统。',
        'capabilities': ['底锁测速', '悬停定点', '相对位移测量']
    },
    'usbl_positioning_system': {
        'name': 'USBL定位系统',
        'aliases': ['USBL定位'],
        'description': '利用水下声学换能器与应答器进行测量，提供机器人相对于母船的绝对三维坐标定位系统。',
        'capabilities': ['水下绝对定位', '目标追踪', '母船相对定位']
    },
    'crawler_module': {
        'name': '履带模块',
        'aliases': ['海床履带模块'],
        'description': '提供海底底质作业行走履带机构，增强机器人在软泥海床的沉陷抵抗力与平稳行走能力。',
        'capabilities': ['海床行走', '稳定支撑', '底质承载']
    },
    'rotary_cleaning_brush': {
        'name': '旋转清洗刷',
        'aliases': ['液压旋转刷'],
        'description': '由液压或电力驱动的高速旋转清理刷头，用于刷除管道接口、阀门控制面板上的海生物与泥沙。',
        'capabilities': ['接口清洁', '检测面清洁', '海生物清理']
    },
    'manipulator_intermeshing_gripper': {
        'name': '机械臂末端夹爪 / 夹具',
        'aliases': ['末端夹爪夹具'],
        'description': '安装于机械臂腕部末端，用于夹持、抓取、持握手柄及固定水下工具工装。',
        'capabilities': ['夹持', '抓取', '操作辅助']
    },
    'hydraulic_cutting_saw_tool': {
        'name': '液压剪切 / 水下锯切工具',
        'aliases': ['水下锯切工具'],
        'description': '高压液压驱动的切削与锯切工装，用于切割构件、水下钢丝绳及清理结构缠绕物。',
        'capabilities': ['构件切割', '线缆切割', '水下锯切']
    },
    'hydraulic_cable_cutter': {
        'name': '液压电缆剪切器',
        'aliases': ['海缆剪切器'],
        'description': '专用于水下强力切断海底电缆、铠装缆绳及管缆防护层的重型液压剪切器。',
        'capabilities': ['海缆剪切', '电缆切断', '应急剪切']
    },
    'electromagnetic_detection_sensor': {
        'name': '电磁检测传感器',
        'aliases': ['电磁异常传感器'],
        'description': '利用电磁感应原理检测金属管道防腐层缺陷、开裂及埋深电磁异常的传感器。',
        'capabilities': ['电磁异常检测', '缺陷检测', '电磁感应']
    },
    'acoustic_transponder': {
        'name': '声学应答器',
        'aliases': ['水下定位应答器'],
        'description': '用于水下目标标定与测距应答的声学设备，向定位系统发送高精度定位脉冲。',
        'capabilities': ['声学测距', '应答定位', '目标标定']
    },
    'light_multi_function_hydraulic_arm': {
        'name': '轻型多功能液压机械臂',
        'aliases': ['轻型7功能臂'],
        'description': '适用于轻型工作级 ROV 的轻量化多功能液压机械臂，用于辅助抓持与精细干预。',
        'capabilities': ['辅助操作', '持物', '水下干预']
    },
    'light_electro_hydraulic_arm': {
        'name': '轻型电液机械臂',
        'aliases': ['轻型电液臂'],
        'description': '适用于轻型作业机器人的电液混合驱动机械臂，提供稳定的辅助抓取与面板操作能力。',
        'capabilities': ['轻型替换操作', '抓取', '阀门干预']
    },
    'depth_gauge': {
        'name': '深度计',
        'aliases': ['水深计'],
        'description': '通过水压感应高精度测量机器人所在水深，支持自动定深悬浮作业。',
        'capabilities': ['深度测量', '定深控制']
    },
    'altimeter': {
        'name': '高度计',
        'aliases': ['离底高度计'],
        'description': '用于高精度测量机器人离海床底部高度，支持离底定高行走与悬浮作业。',
        'capabilities': ['离底高度测量', '定高控制']
    },
    'turbid_water_camera': {
        'name': '浑水摄像机',
        'aliases': ['浑水相机', 'OrphieCam', '浊度透视摄像机'],
        'description': '专用于低能见度和高浑浊度水下环境的强穿透高清晰度光学观察与巡检摄像机。',
        'capabilities': ['低能见度巡检', '高浑浊度光学成像', '水下观测'],
        'weight_in_air_kg': 4.5,
        'weight_in_water_kg': None
    },
    'independent_sampler': {
        'name': '独立采样器',
        'aliases': ['车载独立采样器', '水下独立采样装置'],
        'description': '非机械臂依赖的独立式深海水体与沉积物/环境DNA采样装置。',
        'capabilities': ['独立水体采样', '环境DNA采样', '物理化学采样'],
        'weight_in_air_kg': 15.0,
        'weight_in_water_kg': None
    },
    'valve_torque_tool': {
        'name': '阀门扭矩工具',
        'aliases': ['Torque Tool', '阀门操作工具', 'Class 1-4 Torque Tool'],
        'description': '符合 ISO 13628-8 / API 17H 标准的工作级 ROV 阀门扭矩操作工具，用于水下采油树控制面板及管道阀门的高扭矩旋转开关操作。',
        'capabilities': ['阀门扭矩旋转', 'ISO接口操作', '高扭矩开关'],
        'weight_in_air_kg': 45.0,
        'weight_in_water_kg': 35.0
    },
    'rov_hydraulic_hot_stab': {
        'name': 'ROV液压热插拔接头',
        'aliases': ['Hot Stab', '液压热插拔', '水下热插拔接头'],
        'description': 'ROV 水下液压热插拔接头，用于在水下压力环境下进行流体注液、液压回路高压连接及临时控制通道对接。',
        'capabilities': ['水下液压热插拔', '高压流体连接', '临时液压控制'],
        'weight_in_air_kg': 4.6,
        'weight_in_water_kg': None
    },
    'hydraulic_flying_lead_tool': {
        'name': '液压飞线插拔工具',
        'aliases': ['FLOT', '飞线工具', 'Flying Lead Orientation Tool'],
        'description': '液压飞线定向与对准插拔工具（FLOT），用于辅助 ROV 进行水下液压飞线的精准对准、旋转锁紧与快捷插拔操作。',
        'capabilities': ['液压飞线对准', '飞线旋转插拔', '水下连接干预'],
        'weight_in_air_kg': 45.0,
        'weight_in_water_kg': 25.0
    },
    'electric_flying_lead_tool': {
        'name': '电气飞线插拔工具',
        'aliases': ['EFL插拔工具', '电气飞线工具', 'EFL Connector'],
        'description': '用于水下控制系统电气飞线（EFL）的专用或机械臂直接插拔操作工具，实现水下信号与电力电缆的湿插拔连接。',
        'capabilities': ['电气飞线湿插拔', '信号与电力连接', '水下接头安装'],
        'weight_in_air_kg': 5.0,
        'weight_in_water_kg': None
    },
    'six_axis_force_torque_sensor': {
        'name': '六维力 / 力矩传感器',
        'aliases': ['六维力传感器', '力矩传感器', '6轴力传感器'],
        'description': '安装于机械臂腕部或末端工具接口，实时感知三维空间内的力与力矩大小，为机械臂柔顺控制与精准触觉反馈提供保障。',
        'capabilities': ['六维力控反馈', '触觉感知', '柔顺干预控制'],
        'weight_in_air_kg': 2.0,
        'weight_in_water_kg': None
    },
    'manipulator_wrist_camera': {
        'name': '机械臂腕部摄像机',
        'aliases': ['腕部摄像机', '腕部相机', 'Wrist Camera'],
        'description': '紧凑型深水高清摄像机，直接安装于机械臂腕部，为末端夹爪工具对准、热插拔及精细微操作提供近距离视角。',
        'capabilities': ['末端对准视角', '微操作监控', '近距离光学观察'],
        'weight_in_air_kg': 0.7,
        'weight_in_water_kg': None
    },
    'underwater_imaging_system': {
        'name': '水下成像系统',
        'aliases': ['成像系统', '光学成像系统'],
        'description': '综合水下成像系统，集成高分辨率摄像机、补光灯与画面实时处理模块，用于全过程视觉观察与作业记录。',
        'capabilities': ['全过程视觉观察', '作业画面记录', '水下图像感知'],
        'weight_in_air_kg': None,
        'weight_in_water_kg': None
    },
    'five_dof_manipulator': {
        'name': '五自由度机械臂',
        'aliases': ['5自由度机械臂', '5自由度抓持臂'],
        'description': '5自由度重型抓持或辅助机械臂，主要用于稳定贴靠构件、辅助持物及配合主臂进行水下作业。',
        'capabilities': ['辅助抓持', '构件固定', '辅助干预'],
        'weight_in_air_kg': None,
        'weight_in_water_kg': None
    },
    'manipulator_tool_quick_change': {
        'name': '机械臂工具快换装置',
        'aliases': ['工具快换装置', '快换接头'],
        'description': '用于机械臂末端快速切换不同作业工具的快捷机械/液压锁紧接口。',
        'capabilities': ['末端工具快换', '水下工装切换', '多功能扩展'],
        'weight_in_air_kg': None,
        'weight_in_water_kg': None
    },
    'manipulator_sampling_tool': {
        'name': '机械臂相关采样工具',
        'aliases': ['机械臂采样工具', '机械臂采样工装'],
        'description': '配合机械臂使用的推管采样器、抓斗或防污染密封箱，用于采集深海沉积物、岩石或生物样品。',
        'capabilities': ['机械臂辅助采样', '沉积物/样品采集', '防污染封装'],
        'weight_in_air_kg': None,
        'weight_in_water_kg': None
    },
    'cable_depressor': {
        'name': '压缆器',
        'aliases': ['缆压器', '海缆压块'],
        'description': '配合管缆埋设机使用的压缆引导装置，用于在喷冲开沟后将海底管缆顺畅压入泥面以下并保持埋设姿态。',
        'capabilities': ['管缆压入泥面', '埋设姿态控制', '海缆引导'],
        'weight_in_air_kg': 1000.0,
        'weight_in_water_kg': None
    },
    'burial_depth_sensor': {
        'name': '埋深感应装置',
        'aliases': ['埋深传感器', '埋深检测器'],
        'description': '用于实时感应管缆埋入泥面以下深度的电磁或机械触感测头，配合埋深控制系统进行闭环调节。',
        'capabilities': ['埋深实时感应', '闭环反馈', '路由埋深测量'],
        'weight_in_air_kg': 150.0,
        'weight_in_water_kg': None
    },
    'usbl_auxiliary_positioning_module': {
        'name': 'USBL辅助定位模块',
        'aliases': ['USBL辅助定位', 'AUV定位辅助模块'],
        'description': '专用于 AUV 或小型 ROV 的超短基线辅助声学定位与应答模块，提升定位精度与数据传输效率。',
        'capabilities': ['AUV辅助定位', '声学坐标修正', '水下定位导航'],
        'weight_in_air_kg': None,
        'weight_in_water_kg': None
    },
    'dissolved_oxygen_sensor': {
        'name': '溶解氧传感器',
        'aliases': ['DO传感器', '光学溶解氧传感器'],
        'description': '用于精确测量水体中溶解氧浓度与饱和度的光学/电化学传感器，广泛应用于海洋生态与环境监测。',
        'capabilities': ['溶解氧测量', '海洋水质监测', '生态环境评估'],
        'weight_in_air_kg': None,
        'weight_in_water_kg': None
    },
    'turbidity_sensor': {
        'name': '浊度传感器',
        'aliases': ['光学浊度传感器', '浊度计'],
        'description': '利用光学散射原理测量水体悬浮物浓度与浊度的传感器，用于监测海底施工羽流与泥沙扰动。',
        'capabilities': ['浊度监测', '施工羽流跟踪', '悬浮物浓度测量'],
        'weight_in_air_kg': None,
        'weight_in_water_kg': None
    }
}

for item_key, item_val in distinct_sheet_entries.items():
    catalog[item_key] = item_val

# Clean up aliases so no alias is equal to any catalog item's primary name
all_names = {v['name'] for v in catalog.values()}
for c_key, c_val in catalog.items():
    aliases = c_val.get('aliases', [])
    cleaned = [a for a in aliases if a not in all_names or a == c_val['name']]
    c_val['aliases'] = cleaned

assets['payload_catalog'] = catalog

# Save back to assets.yaml
class CleanDumper(yaml.Dumper):
    def increase_indent(self, flow=False, indentless=False):
        return super(CleanDumper, self).increase_indent(flow=False, indentless=False)

with open(assets_path, 'w', encoding='utf-8') as f:
    yaml.dump(assets, f, Dumper=CleanDumper, allow_unicode=True, sort_keys=False, default_flow_style=False)

print("build_perfect_assets complete!")
